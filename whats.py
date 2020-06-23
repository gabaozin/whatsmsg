"""
Disclaimer: This script is just for educational purpose
"""

#Here is the list of all modules needed for the code to work.
#You can install selenium and openpyxl using pip.

from time import sleep
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains


#Here's how to pull name info from a spreadsheet. Specify PATH to your .XLSX file and put the exact Sheet name.

#In the file, put full names on the A column. Notice that if a cell has been used, it will give out wrong information here. The solution is to select cells below the desired results and click Clear all, found on the Home tab of Microsoft Excel application, or corresponding button in Google Spreadsheet, Libre Office and similar alternatives.

wb = load_workbook("PATH_to_file")
sheet = wb["Sheet tab name"]

names = []

#In case you want to write your message in the Spreadsheet, below is the way to pull the info from there (commented out as we preferred the message to be written in the code)

#str = wb.active["B1"].value
#str2 = wb.active["B2"].value

#Getting all the updated names and appending to the empty names list above.
tuple(sheet['A1':"A" + str(sheet.max_row)])

for rowOfCellObjects in sheet['A1':"A" + str(sheet.max_row)]:
    for cellObj in rowOfCellObjects:
        names.append(cellObj.value)

#in case you want to test the result, you could uncomment the line below
#print(names)

#To get only First names
name = [i.split()[0] for i in names]

#Personalised message, change at will.
#Template = Hi + first name + rest of message (multi-line)

string = 'Hi '
string2 = ''', is everything allright?

I hope you are doing great!

Sending this message to light up your day.

Cheers!!'''

#Calls browser (using Chrome but you can choose browser at will. See selenium documentation), web whatsapp page and then waits on your input(ok) to continue the program.
driver = webdriver.Chrome('PATH_TO_CHROMEDRIVER')
driver.get('https://web.whatsapp.com')
input("Click when QR code done")

#Rapid Tables Notepad is a good page to test changes in the code without bothering people. Uncomment lines 58, 62 and 65 and comment lines 53, 54, 61 and 64 to use this other website. Do the opposite to come back to whatsapp.

#driver.get('https://www.rapidtables.com/tools/notepad.html')

for i in range(len(names)):
    search_box = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, '//*[@id="side"]/div[1]/div/label/div/div[2]')))
    #search_box = driver.find_element_by_id('area')
    search_box.send_keys(names[i] + Keys.ENTER)
    input_box = driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')
    #input_box = driver.find_element_by_id("area")
    input_box.send_keys(string + name[i])
    for line in string2.split('\n'):
        ActionChains(driver).send_keys(line).perform()
        ActionChains(driver).key_down(Keys.SHIFT).key_down(Keys.ENTER).key_up(Keys.SHIFT).key_up(Keys.ENTER).perform()
    ActionChains(driver).send_keys(Keys.RETURN).perform()
